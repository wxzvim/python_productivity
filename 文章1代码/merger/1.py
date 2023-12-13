#!/usr/bin/python3
import openpyxl
import os


from pathlib import Path, PurePath

src_path = "../调查问卷/"
des_path = "../results/results.xlsx"
#
p = Path(src_path)
files = [x for x in p.iterdir() if PurePath(x).match("*.xlsx")]

content = []

for file in files:
    username = file.stem
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    cell_value1 = sheet["E5"].value
    cell_value2 = sheet["E11"].value
    temp = f"{username},{cell_value1},{cell_value2}"
    content.append(temp.split(","))
    print(temp)


table_header = ["员工姓名", "第一题", "第二题"]

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "统计结果"

row_num = 1


for col_num, header in enumerate(table_header, start=1):
    sheet.cell(row=row_num, column=col_num, value=header)

row_num += 1


for line in content:
    col_num = 1
    for cell in line:
        sheet.cell(row=row_num, column=col_num, value=cell)
        col_num += 1

    row_num += 1


os.makedirs(os.path.dirname(des_path), exist_ok=True)

workbook.save(des_path)
