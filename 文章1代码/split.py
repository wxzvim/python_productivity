import openpyxl
from pathlib import Path
import os

from openpyxl.compat import numbers


src_file = "./工资单/salary-list.xlsx"
dest_path = Path("./results")
print(os.getcwd())

try:
    workbook = openpyxl.load_workbook(src_file)
    print("successful")
except openpyxl.utils.exceptions.InvalidFileException:
    print("fail: format not suppose and file bad")
except FileNotFoundError:
    print("file not find")
except Exception as e:
    print(f"other error: {e}")
else:
    sheet = workbook.active

cell_headers = [cell.value for cell in sheet[1]]
print(cell_headers)


def write_to_file(filename, content):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.title = "本月工资"

    row = 1
    for line in content:
        col = 1
        for cell in line:
            sheet.cell(row=row, column=col, value=cell)
            col += 1
        row += 1
    workbook.save("filename.xlsx")


number_row = sheet.max_row

for line in range(2, number_row + 1):
    content = [cell.value for cell in sheet[line]]
    new_content = []
    new_content.append(cell_headers)
    new_content.append(content)
    write_to_file(filename=content[2], content=new_content)
